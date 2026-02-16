"""
Excel / CSV Document Loader
==============================
Ingests .xlsx, .xls, and .csv files into the RAG pipeline.

Common telecom use cases:
  - Rate decks
  - Billing summaries
  - Revenue reports
  - Contact directories
  - Regulatory filing data

Pipeline:
  1. Detect file type
  2. Load with openpyxl (xlsx) / xlrd (xls) / csv
  3. Convert rows to text chunks
  4. Optionally: detect headers and create structured records
"""

import os
import csv
import io
from typing import List, Optional, Dict, Any
from langchain_core.documents import Document


class ExcelCSVLoader:
    """
    Loads Excel and CSV files into Document objects.
    Each sheet → separate documents, tables → text chunks.
    """

    def load(self, file_path: str) -> List[Document]:
        """
        Load an Excel, CSV, or TSV file.

        Args:
            file_path: Path to .xlsx, .xls, .csv, or .tsv file

        Returns:
            List of Document objects
        """
        ext = file_path.rsplit(".", 1)[-1].lower() if "." in file_path else ""
        base_name = os.path.basename(file_path)

        if ext in ("csv", "tsv"):
            return self._load_csv(file_path, base_name, delimiter="\t" if ext == "tsv" else None)
        elif ext in ("xlsx", "xls"):
            return self._load_excel(file_path, base_name)
        else:
            print(f"[EXCEL] Unsupported file type: {ext}")
            return []

    def _load_csv(self, file_path: str, source: str, delimiter: Optional[str] = None) -> List[Document]:
        """Load a CSV or TSV file."""
        documents = []

        try:
            # Detect encoding
            encoding = "utf-8"
            try:
                with open(file_path, "r", encoding="utf-8") as f:
                    f.read(1024)
            except UnicodeDecodeError:
                encoding = "latin-1"

            with open(file_path, "r", encoding=encoding, newline="") as f:
                # Use provided delimiter or auto-detect
                if delimiter is None:
                    sample = f.read(4096)
                    f.seek(0)
                    try:
                        dialect = csv.Sniffer().sniff(sample)
                        delimiter = dialect.delimiter
                    except csv.Error:
                        delimiter = ","

                reader = csv.reader(f, delimiter=delimiter)
                rows = list(reader)

            if not rows:
                return []

            # First row as headers
            headers = rows[0]
            data_rows = rows[1:]

            # Create a summary document
            summary_lines = [f"CSV File: {source}"]
            summary_lines.append(f"Columns: {', '.join(headers)}")
            summary_lines.append(f"Total Rows: {len(data_rows)}")
            summary_lines.append("")

            documents.append(Document(
                page_content="\n".join(summary_lines),
                metadata={
                    "source": source,
                    "page": 0,
                    "extraction_method": "csv_loader",
                    "element_type": "table_summary",
                    "row_count": len(data_rows),
                    "column_count": len(headers),
                },
            ))

            # Chunk rows into groups (25 rows per chunk)
            chunk_size = 25
            for i in range(0, len(data_rows), chunk_size):
                chunk_rows = data_rows[i:i + chunk_size]
                text_lines = []
                for row in chunk_rows:
                    pairs = []
                    for j, cell in enumerate(row):
                        header = headers[j] if j < len(headers) else f"Column_{j}"
                        if cell and cell.strip():
                            pairs.append(f"{header}: {cell.strip()}")
                    if pairs:
                        text_lines.append(" | ".join(pairs))

                if text_lines:
                    documents.append(Document(
                        page_content="\n".join(text_lines),
                        metadata={
                            "source": source,
                            "page": (i // chunk_size) + 1,
                            "extraction_method": "csv_loader",
                            "element_type": "table_data",
                            "row_start": i + 1,
                            "row_end": min(i + chunk_size, len(data_rows)),
                        },
                    ))

            print(f"[EXCEL] Loaded CSV: {len(data_rows)} rows, {len(documents)} chunks")

        except Exception as e:
            print(f"[EXCEL] CSV loading failed: {e}")

        return documents

    def _load_excel(self, file_path: str, source: str) -> List[Document]:
        """Load an Excel file (xlsx/xls)."""
        documents = []

        try:
            import openpyxl

            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    rows.append([str(cell) if cell is not None else "" for cell in row])

                if not rows:
                    continue

                # First row as headers
                headers = rows[0]
                data_rows = rows[1:]

                # Summary
                summary = f"Excel Sheet: {sheet_name} (from {source})\n"
                summary += f"Columns: {', '.join(h for h in headers if h)}\n"
                summary += f"Total Rows: {len(data_rows)}\n"

                documents.append(Document(
                    page_content=summary,
                    metadata={
                        "source": source,
                        "page": 0,
                        "sheet": sheet_name,
                        "extraction_method": "excel_loader",
                        "element_type": "table_summary",
                        "row_count": len(data_rows),
                    },
                ))

                # Chunk data rows
                chunk_size = 25
                for i in range(0, len(data_rows), chunk_size):
                    chunk_rows = data_rows[i:i + chunk_size]
                    text_lines = []
                    for row in chunk_rows:
                        pairs = []
                        for j, cell in enumerate(row):
                            header = headers[j] if j < len(headers) else f"Column_{j}"
                            if cell and cell.strip():
                                pairs.append(f"{header}: {cell.strip()}")
                        if pairs:
                            text_lines.append(" | ".join(pairs))

                    if text_lines:
                        documents.append(Document(
                            page_content="\n".join(text_lines),
                            metadata={
                                "source": source,
                                "page": (i // chunk_size) + 1,
                                "sheet": sheet_name,
                                "extraction_method": "excel_loader",
                                "element_type": "table_data",
                                "row_start": i + 1,
                                "row_end": min(i + chunk_size, len(data_rows)),
                            },
                        ))

                print(f"[EXCEL] Loaded sheet '{sheet_name}': {len(data_rows)} rows")

            wb.close()

        except ImportError:
            print("[EXCEL] openpyxl not installed. pip install openpyxl")
            # Try pandas as fallback
            return self._load_excel_pandas(file_path, source)
        except Exception as e:
            print(f"[EXCEL] Excel loading failed: {e}")
            return self._load_excel_pandas(file_path, source)

        return documents

    def _load_excel_pandas(self, file_path: str, source: str) -> List[Document]:
        """Fallback: load Excel using pandas."""
        documents = []

        try:
            import pandas as pd

            ext = file_path.rsplit(".", 1)[-1].lower()
            engine = "openpyxl" if ext == "xlsx" else "xlrd"

            xls = pd.ExcelFile(file_path, engine=engine)

            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name)

                if df.empty:
                    continue

                # Convert to text
                text = f"Sheet: {sheet_name}\n"
                text += df.to_string(index=False, max_rows=100)

                documents.append(Document(
                    page_content=text,
                    metadata={
                        "source": source,
                        "page": 1,
                        "sheet": sheet_name,
                        "extraction_method": "pandas_excel",
                        "element_type": "table_data",
                        "row_count": len(df),
                    },
                ))

            print(f"[EXCEL] Loaded {len(documents)} sheets via pandas fallback")

        except Exception as e:
            print(f"[EXCEL] Pandas fallback also failed: {e}")

        return documents


# Module-level singleton
_excel_loader: Optional[ExcelCSVLoader] = None


def get_excel_loader() -> ExcelCSVLoader:
    global _excel_loader
    if _excel_loader is None:
        _excel_loader = ExcelCSVLoader()
    return _excel_loader

